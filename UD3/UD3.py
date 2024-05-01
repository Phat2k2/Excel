import numpy as np
from numpy import linalg as LA
from numpy.linalg import inv
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import math
import random
import os
# chon cac gia tri n,r trong thuat toan
n = 2
r = 1
# Ma tran A
A = np.array([[0,0],[0,-8]])
# tinh cac tri rieng cua ma tran A, w la danh sach cac tri rieng, v la ma tran cheo hoa
w,v=LA.eig(A)
# chon rho > 0 trong thuat toan
p = 1
# vecto b
b = np.array([1,0])
# ma tran don vi trong R^n+
I = np.identity(n)
res = np.array([[-1/8,math.sqrt(63)/8],[-1/8,-math.sqrt(63)/8], [-1,0]])
count= [0,0,0]
# cong thuc tim hinh chieu
def hc(x):
  if LA.norm(x) > r:
    x = r*x/LA.norm(x)
  return x

wb = load_workbook('UD3.xlsx')
ws = wb.active


# Thuat toan DC
for i in range(100000):
    def er(x,k):
        if LA.norm(x[k%2]) > 1:
            return LA.norm(x[(k+1)%2] - x[k%2])/LA.norm(x[k%2])
        else:
            return LA.norm(x[(k+1)%2] - x[k%2])
    k = 0
    print(i)
    X = [ np.random.randint(-1000,1001,size = (n) ), np.random.randint(-100,100, size = (n)) ]
    while er(X,k) >= 1e-6:
        X[(k+1)%2] = hc(  1/p*np.dot(X[k%2], p*I - A) - b/p)
        k = k + 1
    ws['A' + str(i+2)]  = i + 1
    ws['B' + str(i+2)]  = '(' + str(X[k%2][0]) + ',' + str(X[k%2][1]) +  ')'
    ws['C' + str(i+2)]  = min(LA.norm(X[k%2] -res[0]),LA.norm(X[k%2] -res[1]),LA.norm(X[k%2] -res[2]))
    for j in range(3):
        if  ws['C' + str(i+2)].value == LA.norm(X[k%2] -res[j]):
            ws['D' + str(i+2)] = j + 1
            count[j] = count[j] + 1
ws['G' + str(2)] = count[0]
ws['H' + str(2)] = count[1]
ws['I' + str(2)] = count[2]
wb.save('UD3.xlsx')
try:
    os.startfile("UD3.xlsx")
except OSError as e:
    print("Không thể mở file:", e)

