import numpy as np
from numpy import linalg as LA
from numpy.linalg import inv
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import math
import random
import os
# chon cac gia tri n,r trong thuat toan
n = 2
r = 1
# Ma tran A
A = np.array([[0,0],[0,1]])
# tinh cac tri rieng cua ma tran A, w la danh sach cac tri rieng, v la ma tran cheo hoa
w,v=LA.eig(A)
# chon rho > 0 trong thuat toan
p = 1
# vecto b
b = np.array([0,1])
# ma tran don vi trong R^n+
I = np.identity(n)

# cong thuc tim hinh chieu
def hc(x):
  if LA.norm(x) > r:
    x = r*x/LA.norm(x)
  return x
wb = load_workbook('UD1.xlsx')
ws = wb.active

# Thuat toan DC
for i in range(1000):
    k = 0
    def er(x,k):
        if LA.norm(x[k%2]) > 1:
            return LA.norm(x[(k+1)%2] - x[k%2])/LA.norm(x[k%2])
        else:
            return LA.norm(x[(k+1)%2] - x[k%2])
    X = [ np.random.uniform(-1000,1000,size = (n) ), np.random.uniform(-1000,1000, size = (n)) ]
    print(i)
    while er(X,k) >= 1e-6:
        X[(k+1)%2] = hc(  1/p*np.dot(X[k%2], p*I - A) - b/p)
        k = k + 1
    ws['A' + str(i+2)]  = i + 1
    ws['B' + str(i+2)]  = '(' + str(X[k%2][0]) + ',' + str(X[k%2][1]) +  ')'
    ws['C' + str(i+2)]  = LA.norm(X[k%2] - np.array([0,-1]))
wb.save('UD1.xlsx')
try:
    os.startfile("UD1.xlsx")
except OSError as e:
    print("Không thể mở file:", e)
