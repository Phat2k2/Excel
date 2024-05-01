import numpy as np
from numpy import linalg as LA
from numpy.linalg import inv
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import math
import random
import os

def rrr(n,h):
    k = 0
    arr = [0,0,0,0]
    for i in range(h):
        print(k)
        res = rr(n)
        for j in range(4):
             arr[j] = arr[j] + res[j]
        k = k + 1
    for j in range(4):
        arr[j] = arr[j]/h
    return arr
def gen_matrix(n,num):
    I = np.identity(n)
    U = I
    for i in range(3):
        w = np.random.uniform(-1,1,n)
        Q = I - 2*np.outer(w,w.T)/LA.norm(w)**2
        U = U.dot(Q)
    g = np.random.uniform(-1,1,n)
    b = U.dot(g)
    if num==1:
        d = np.random.uniform(0,5,n)
        while min(d) == 0:
            d = np.random.uniform(0,5,n)        
    else:
        d = np.random.uniform(-5,5,n)
        while min(d) > 0:
            d = np.random.uniform(-5,5,n)            
    A = U.dot(np.diag(d)).dot(U.T)
    return [A,b]

def rr(n):
    A1, b1 = gen_matrix(n,1)
    A2, b2 = gen_matrix(n,0)
    b = (b1+b2)/2
    # vecto b
    # ma tran don vi trong R^n+
    I = np.identity(n)
    r = np.random.uniform(1,100)
    # cong thuc tim hinh chieu
    def hc(x):
      if LA.norm(x) > r:
        x = r*x/LA.norm(x)
      return x

    def er(x,k):
        if LA.norm(x[k%2]) > 1:
            return LA.norm(x[(k+1)%2] - x[k%2])/LA.norm(x[k%2])
        else:
            return LA.norm(x[(k+1)%2] - x[k%2])

    start_time1 = time.time()
    k1 = 0
    w,v=LA.eig(A1)
    p1 = abs(max(w)) + 1        
    X0 = np.tile(r/math.sqrt(n), n)
    X = np.array([X0, np.tile(0, n)])    
    while er(X,k1) >= 1e-6:
        X[(k1+1)%2] = hc(  1/p1*np.dot(X[k1%2], p1*I - A1) - b/p1)
        k1 = k1 + 1
    end_time1 = time.time()
    
    start_time2 = time.time()
    w,v=LA.eig(A2)
    p2 = abs(max(w)) + 1        
    X0 = np.tile(r/math.sqrt(n), n)
    X = np.array([X0, np.tile(0, n)])
    k2 = 0
    while er(X,k2) >= 1e-6:
        X[(k2+1)%2] = hc(  1/p2*np.dot(X[k2%2], p2*I - A2) - b/p2)
        k2 = k2 + 1
    end_time2 = time.time()
    return [k1,k2, end_time1-start_time1,  end_time2-start_time2 ]
    

file_name = '21-04-2024.xlsx'
wb = load_workbook(file_name)
ws = wb.active
i = 3
print('A' + str(i))
while ws['A' + str(i)].value is not None:
    print(i,ws['A' + str(i)].value)
    n = ws['A'+ str(i)].value
    arr = rrr(n,100)
    ws['B' + str(i)]  = arr[0]
    ws['C' + str(i)] = arr[1]
    ws['D' + str(i)] = arr[2]
    ws['E' + str(i)] = arr[3] 
    i = i + 1
    wb.save(file_name)  
try:
    os.startfile(file_name)
except OSError as e:
    print("Không thể mở file:", e)